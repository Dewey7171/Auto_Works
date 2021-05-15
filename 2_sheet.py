from openpyxl import Workbook
wb = Workbook()

# wb.active
ws = wb.create_sheet() #새로운 Sheet 기본 이름으로 생성 

ws.title = "SJsheet" # Sheet이름 변경

ws.sheet_properties.tabColor = "bd7ff7" # RGB형태로 값을 넣어주면 탭 색상이 변경됨

#Sheet, SJSheet, NextSheet
ws1 = wb.create_sheet("NextSheet") # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet",2) # 2번째 index에 Sheet 생성

new_ws = wb["NewSheet"] #Dict 형태로 Sheet 접근 가능

print(wb. sheetnames) # 모든 sheet 이름 확인

# Sheet 복사 
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copued Sheet"

wb.save("sample.xlsx")

