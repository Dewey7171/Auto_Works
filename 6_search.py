from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws= wb.active

# 영어 항목에 있는 숫자크기가 50 이상이면 출력함

for row in ws.iter_rows(min_row=2):
    #번호, 영어, 수학
    if int(row[1].value) > 50:
        print(row[0].value,"번 학생은 영어 잘함")


#영어 항목의 이름을 컴퓨터로 바꾸는 방법 

for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"

wb.save("sample1.xlsx")