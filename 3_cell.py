from openpyxl import Workbook
wb = Workbook()

ws = wb.active
ws.title = "SJsheet"

# A1 셀에 1 이라는 값을 입력 
ws["A1"] = 3
ws["A2"] = "천재"
ws["A3"] = "대단해"

ws["B1"] = "하이"
ws["B2"] = "dd"
ws["B3"] = "gg"

print(ws["A1"]) #A1 셀의 정보 출력 
print(ws["A1"].value) # A1 셀의 값을 출력
print(ws["A10"].value) # 값이 없으면 none 출력

# row = 1, 2, 3, ...
# column = A(1),B(2),C(3) ...

print(ws.cell(row =1 , column=1).value) #ws ["A1"].value

print(ws.cell(row =1, column=2).value) #ws ["B1"].value

c = ws.cell(column=3, row=1, value = 10) #ws["C1"].value = 10
print(c.value)


from random import *
index = 1

#반복문을 이용해 랜덤 숫자 채우기 
for x in range(1,11): #10개 row 
    for y in range (1,11): #10개 column
        #ws.cell(row =x , column = y, value=randint(0,100))
        ws.cell(row=x, column =y, value=index)
        index += 1





wb.save("sample.xlsx")