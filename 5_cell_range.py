from openpyxl import Workbook
from random import *
wb = Workbook()
ws= wb.active

#한 줄씩 데이터 넣기 
ws.append(["번호", "영어", "수학"]) # A, B, C

for i in range(1,11):
    ws.append([i,randint(0,100),randint(0,100) ])



col_B = ws["B"] #영어 Column만 가져옴 

# for cell in col_B:
#     print(cell.value)

# col_range = ws["B:C"] #영어 수학 column 같이 가져옴
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

row_title = ws[1] #1번째 row만 가져오기 

# for cell in row_title:
#     print(cell.value)

# row_range = ws[2:6] #1번 줄 title 빼고 2부터 6번쨰 줄까지 가져옴 

# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end = " ")
#     print()

# from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row]

# for rows in row_range:
#     for cell in rows:
#         #print(cell.value, end=" ")
#         #print(cell.coordinate, end = " ")
#         xy = coordinate_from_string(cell.coordinate)
#         # print(xy, end= " ")
#         print(xy[0], end = "") # A
#         print(xy[1], end = " ") # 1

#     print()

#전체 rows
# print(tuple(ws.rows))
# for row in tuple(ws.rows):
#     print(row[2].value)


# 전체 columns
#print(tuple(ws.columns))
# for column in tuple(ws.columns):
#     print(column[0].value)

# for row in ws.iter_rows(min_row = 2, max_row=11, min_col=2, max_col=3): #전체 row 
#     #print(row[0].value, row[1].value)
#     print(row)

for col in ws.iter_cols(min_row = 1, max_row=5, min_col=1, max_col=3):
    print(col)
# for column in ws.iter_cols():
#     print(column[0].value)

#row는 가로 줄을 가져온다 
#col은 세로 줄을 가져온다.

wb.save("sample.xlsx")